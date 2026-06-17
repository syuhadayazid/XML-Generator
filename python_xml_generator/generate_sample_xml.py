#!/usr/bin/env python3
"""Generate a sample XML document from an Excel mapping file.

Expected Excel shape:
- One column containing XPath-like destination paths.
- Optional second column containing mapping guidance or explicit values.

Example path:
  /px:NotifyShipment/px:DataArea/px:Shipment/px:ShipmentHeader/px:ID[@typeCode='ABC']
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime dependency guard
    print(
        "Missing dependency: openpyxl. Install it with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


NO_MAPPING_TOKEN = "__NO_MAPPING__"


def normalize_path(path_text: str) -> str:
    text = (path_text or "").strip()
    if not text:
        return ""

    if not text.startswith("/"):
        text = "/" + text

    while "px:px:" in text:
        text = text.replace("px:px:", "px:")

    text = text.replace("px:/", "px:")
    text = text.replace("//", "/")
    text = re.sub(r"(?<=[A-Za-z0-9_\]])(px:)(?=[A-Za-z_])", r"/\1", text)
    text = re.sub(r"(\[[^\]]+\])(?=[A-Za-z_])", r"\1/", text)
    return text


def split_segments(path_text: str) -> list[str]:
    segments: list[str] = []
    current: list[str] = []
    depth = 0
    for ch in path_text:
        if ch == "/" and depth == 0:
            segment = "".join(current).strip()
            if segment:
                segments.append(segment)
            current = []
            continue
        if ch == "[":
            depth += 1
        elif ch == "]" and depth > 0:
            depth -= 1
        current.append(ch)

    tail = "".join(current).strip()
    if tail:
        segments.append(tail)
    return segments


def local_name(name: str) -> str:
    return name.split(":", 1)[1] if ":" in name else name


def parse_segment(segment: str) -> tuple[str, dict[str, str]]:
    base = segment
    predicates = re.findall(r"\[([^\]]+)\]", segment)
    if "[" in segment:
        base = segment.split("[", 1)[0]

    attrs: dict[str, str] = {}
    for pred in predicates:
        pred = pred.strip()
        # Supports forms like @typeCode='X' and /px:@typeCode = "X"
        match = re.match(
            r"^/?(?:[A-Za-z_][\w.-]*:)?@([A-Za-z_][\w.-]*)\s*=\s*['\"]([^'\"]+)['\"]$",
            pred,
        )
        if match:
            attrs[match.group(1)] = match.group(2)

    return base.strip(), attrs


def pick_value(raw_value: str | None, leaf_name: str, force_sample_value: bool = False) -> str | None:
    if force_sample_value:
        return "SAMPLE_VALUE"

    if raw_value is None:
        raw_value = ""
    text = str(raw_value).replace("\r", "").strip()
    if not text:
        if "datetime" in leaf_name.lower():
            return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        return "SAMPLE_VALUE"

    first_line = text.split("\n", 1)[0].strip()
    if not first_line:
        return "SAMPLE_VALUE"

    hardcode_match = re.search(r"(?i)hardcode\s*(?:=|to)?\s*['\"]?([A-Za-z0-9_:\- ]+)['\"]?", first_line)
    if hardcode_match:
        return hardcode_match.group(1).strip()

    if re.fullmatch(r"(?i)no\s+mapping", first_line):
        return NO_MAPPING_TOKEN

    if re.search(r"(?i)current\s+date|system\s+date|today", first_line):
        return dt.datetime.now().strftime("%Y%m%d")

    if re.search(r"(?i)current\s+time|system\s+time|time\s+now", first_line):
        return dt.datetime.now().strftime("%H%M")

    # Guidance text should not become literal XML values.
    if re.search(r"(?i)format|expressed\s+as|for\s+example|if\s+.+\s+then\s+map", first_line):
        if "datetime" in leaf_name.lower():
            return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        return "SAMPLE_VALUE"

    return first_line


def find_or_create_child(parent: ET.Element, tag: str, attrs: dict[str, str]) -> ET.Element:
    for child in list(parent):
        if child.tag != tag:
            continue
        if all(child.attrib.get(k) == v for k, v in attrs.items()):
            return child

    child = ET.SubElement(parent, tag)
    for key, value in attrs.items():
        child.set(key, value)
    return child


def set_element_text(node: ET.Element, value: str) -> None:
    if node.text is None or not node.text.strip() or node.text.strip() == "SAMPLE_VALUE":
        node.text = value


def build_xml(
    rows: Iterable[tuple[str, str | None]],
    force_sample_value: bool = False,
) -> ET.ElementTree:
    root: ET.Element | None = None

    for raw_path, raw_value in rows:
        normalized = normalize_path(raw_path)
        if not normalized:
            continue

        segments = split_segments(normalized)
        if not segments:
            continue

        current: ET.Element | None = None
        last_element_name = ""

        for idx, raw_segment in enumerate(segments):
            seg = raw_segment.strip()
            if not seg:
                continue

            if seg.startswith("@") or seg.startswith("px:@"):
                if current is not None:
                    attr_name = local_name(seg[1:] if seg.startswith("@") else seg[4:])
                    attr_value = pick_value(raw_value, last_element_name, force_sample_value)
                    if attr_value and attr_value != NO_MAPPING_TOKEN:
                        current.set(attr_name, attr_value)
                continue

            element_name, pred_attrs = parse_segment(seg)
            tag = element_name.strip()
            if not tag:
                continue

            last_element_name = local_name(tag)

            if idx == 0:
                if root is None:
                    root = ET.Element(tag)
                    for key, value in pred_attrs.items():
                        root.set(key, value)
                current = root
                continue

            if current is None:
                continue
            current = find_or_create_child(current, tag, pred_attrs)

        if current is not None:
            value = pick_value(raw_value, last_element_name, force_sample_value)
            if value and value != NO_MAPPING_TOKEN and not list(current):
                set_element_text(current, value)

    if root is None:
        root = ET.Element("Root")

    return ET.ElementTree(root)


def excel_col_to_index(column_ref: str) -> int:
    text = (column_ref or "").strip()
    if not text:
        raise ValueError("Empty column reference.")

    if text.isdigit():
        idx = int(text)
        if idx < 1:
            raise ValueError("Column index must be >= 1.")
        return idx - 1

    if not re.fullmatch(r"[A-Za-z]+", text):
        raise ValueError(f"Invalid column reference: {column_ref}")

    result = 0
    for ch in text.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def looks_like_xml_path(value: str) -> bool:
    text = (value or "").strip()
    if not text:
        return False
    if "/" not in text:
        return False
    if text.count("/") < 2:
        return False
    return bool(re.search(r"[A-Za-z_][\w.-]*(:[A-Za-z_][\w.-]*)?", text))


def detect_xml_path_column(rows: list[tuple], start_row: int = 0) -> int | None:
    sample_rows = rows[start_row : start_row + 400]
    if not sample_rows:
        return None

    max_cols = max((len(r) for r in sample_rows), default=0)
    if max_cols == 0:
        return None

    best_col: int | None = None
    best_hits = 0
    best_ratio = 0.0

    for col_idx in range(max_cols):
        non_empty = 0
        hits = 0
        for row in sample_rows:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            non_empty += 1
            if looks_like_xml_path(text):
                hits += 1

        if non_empty == 0:
            continue

        ratio = hits / non_empty
        if hits > best_hits or (hits == best_hits and ratio > best_ratio):
            best_col = col_idx
            best_hits = hits
            best_ratio = ratio

    if best_col is None:
        return None

    # Require enough evidence to avoid random text columns.
    if best_hits < 3 or best_ratio < 0.4:
        return None

    return best_col


def group_same_field_siblings(element: ET.Element) -> None:
    children = list(element)
    if not children:
        return

    grouped: dict[str, list[ET.Element]] = {}
    tag_order: list[str] = []
    for child in children:
        if child.tag not in grouped:
            grouped[child.tag] = []
            tag_order.append(child.tag)
        grouped[child.tag].append(child)

    reordered: list[ET.Element] = []
    for tag in tag_order:
        reordered.extend(grouped[tag])

    element[:] = reordered

    for child in element:
        group_same_field_siblings(child)


def indent_xml(element: ET.Element, level: int = 0) -> None:
    pad = "  "
    indentation = "\n" + level * pad
    if len(element):
        if not element.text or not element.text.strip():
            element.text = indentation + pad
        for child in element:
            indent_xml(child, level + 1)
        if not element[-1].tail or not element[-1].tail.strip():
            element[-1].tail = indentation
    if level and (not element.tail or not element.tail.strip()):
        element.tail = indentation


def load_rows_from_xlsx(
    file_path: Path,
    path_column: str,
    value_column: str,
    worksheet_name: str | None,
    path_column_ref: str | None = None,
    sample_only: bool = False,
) -> list[tuple[str, str | None]]:
    workbook = load_workbook(filename=file_path, data_only=True)

    if worksheet_name:
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{worksheet_name}' not found in workbook.")
        sheet = workbook[worksheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return []

    if path_column_ref:
        path_idx = excel_col_to_index(path_column_ref)
        loaded: list[tuple[str, str | None]] = []
        for row in all_rows:
            path_cell = row[path_idx] if path_idx < len(row) else None
            if path_cell is None:
                continue
            path_text = str(path_cell).strip()
            if not looks_like_xml_path(path_text):
                continue
            loaded.append((path_text, None if sample_only else "SAMPLE_VALUE"))
        return loaded

    normalized_target_path = path_column.strip().lower()
    normalized_fallback_path = "element xpath or segment, loop, element identifier"

    header_row_idx = 0
    header = all_rows[0]
    for idx, candidate in enumerate(all_rows[:50]):
        normalized_cells = {
            str(cell).strip().lower()
            for cell in candidate
            if cell is not None and str(cell).strip()
        }
        if (
            normalized_target_path in normalized_cells
            or normalized_fallback_path in normalized_cells
        ):
            header_row_idx = idx
            header = candidate
            break

    header_map: dict[str, int] = {}
    for idx, name in enumerate(header):
        if name is None:
            continue
        normalized_name = str(name).strip().lower()
        if normalized_name not in header_map:
            header_map[normalized_name] = idx

    path_idx = header_map.get(path_column.strip().lower())
    value_idx = header_map.get(value_column.strip().lower())

    # Auto-detect XML path column when header-based lookup is unavailable.
    if path_idx is None:
        path_idx = detect_xml_path_column(all_rows, start_row=header_row_idx + 1)

    # Fallback for loosely structured files.
    if path_idx is None:
        path_idx = 0
    if value_idx is None:
        value_idx = header_map.get("mapping rules")
    if value_idx is None:
        value_idx = None

    loaded: list[tuple[str, str | None]] = []
    for row in all_rows[header_row_idx + 1 :]:
        path_cell = row[path_idx] if path_idx < len(row) else None
        if path_cell is None:
            continue
        path_text = str(path_cell).strip()
        if not looks_like_xml_path(path_text):
            continue
        value_text = None
        if value_idx is not None and value_idx < len(row):
            cell_value = row[value_idx]
            value_text = None if cell_value is None else str(cell_value)
        loaded.append((path_text, value_text))

    return loaded


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a sample XML file from Excel path/value mapping rows."
    )
    parser.add_argument("input", help="Path to .xlsx mapping file")
    parser.add_argument(
        "-o",
        "--output",
        default="sample.xml",
        help="Output XML file path (default: sample.xml)",
    )
    parser.add_argument(
        "--path-column",
        default="Element Xpath or Segment, Loop, Element Identifier",
        help="Header text for the path column",
    )
    parser.add_argument(
        "--value-column",
        default="Value",
        help="Header text for the value/guidance column",
    )
    parser.add_argument(
        "--worksheet",
        default=None,
        help="Optional worksheet name (default: first sheet)",
    )
    parser.add_argument(
        "--path-column-ref",
        default=None,
        help="Use a fixed Excel column for paths (e.g. B or 2), bypassing header detection",
    )
    parser.add_argument(
        "--sample-only",
        action="store_true",
        help="Force all generated text/attribute values to SAMPLE_VALUE",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if input_path.suffix.lower() != ".xlsx":
        print("Input must be an .xlsx file.", file=sys.stderr)
        return 1

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1

    try:
        rows = load_rows_from_xlsx(
            file_path=input_path,
            path_column=args.path_column,
            value_column=args.value_column,
            worksheet_name=args.worksheet,
            path_column_ref=args.path_column_ref,
            sample_only=args.sample_only,
        )
    except Exception as exc:
        print(f"Failed to read Excel file: {exc}", file=sys.stderr)
        return 1

    if not rows:
        print("No usable mapping rows found in the Excel file.", file=sys.stderr)
        return 1

    tree = build_xml(rows, force_sample_value=args.sample_only)
    group_same_field_siblings(tree.getroot())
    indent_xml(tree.getroot())
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tree.write(output_path, encoding="utf-8", xml_declaration=True)

    print(f"Generated XML file: {output_path}")
    print(f"Rows processed: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
