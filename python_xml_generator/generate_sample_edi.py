#!/usr/bin/env python3
"""Generate a sample X12 EDI document from an Excel mapping file.

Expected input:
- One column containing X12-style paths, for example:
  /X12/TS_214/GROUP_1/B10/B1001
  /X12/ISA/ISA01
- Optional value guidance column (for hardcoded values).
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import NamedTuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime dependency guard
    print(
        "Missing dependency: openpyxl. Install it with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


NO_MAPPING_TOKEN = "__NO_MAPPING__"
CONTROL_SEGMENTS = {"ISA", "GS", "ST", "SE", "GE", "IEA"}


class SefToken(NamedTuple):
    segment_id: str
    is_required: bool
    max_use: int


class SefSegmentDefinition(NamedTuple):
    segment_id: str
    element_count: int
    required_positions: list[int]
    element_ids: list[str]


class SefElementDefinition(NamedTuple):
    element_id: str
    data_type: str
    min_length: int
    max_length: int


def local_name(name: str) -> str:
    return name.split(":", 1)[1] if ":" in name else name


def excel_col_to_index(column_ref: str) -> int:
    text = (column_ref or "").strip()
    if not text:
        raise ValueError("Empty column reference.")

    if text.isdigit():
        idx = int(text)
        if idx < 1:
            raise ValueError("Column index must be >= 1.")
        return idx - 1

    if not re.fullmatch(r"[A-Za-z]+", text):
        raise ValueError(f"Invalid column reference: {column_ref}")

    result = 0
    for ch in text.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def index_to_excel_col(col_idx: int) -> str:
    value = col_idx + 1
    out = ""
    while value > 0:
        rem = (value - 1) % 26
        out = chr(ord("A") + rem) + out
        value = (value - 1) // 26
    return out


def looks_like_x12_path(value: str) -> bool:
    text = (value or "").strip()
    if not text:
        return False
    if "/" not in text:
        return False
    if "X12" not in text.upper():
        return False
    return bool(re.search(r"/[A-Za-z0-9_:-]{2,}/[A-Za-z0-9_:-]{3,}", text))


def normalize_path(path: str) -> str:
    text = (path or "").strip()
    if not text:
        return ""
    if not text.startswith("/"):
        text = "/" + text
    text = re.sub(r"/{2,}", "/", text)
    return text


def split_segments(path: str) -> list[str]:
    segments: list[str] = []
    current: list[str] = []
    depth = 0

    for ch in path:
        if ch == "/" and depth == 0:
            segment = "".join(current).strip()
            if segment:
                segments.append(segment)
            current = []
            continue

        if ch == "[":
            depth += 1
        elif ch == "]" and depth > 0:
            depth -= 1

        current.append(ch)

    tail = "".join(current).strip()
    if tail:
        segments.append(tail)

    return segments


def strip_segment_predicate(segment: str) -> str:
    text = str(segment or "").strip()
    if not text:
        return ""
    if "[" in text:
        return text.split("[", 1)[0].strip()
    return text


def extract_segment_qualifier(segment: str, segment_id: str) -> str | None:
    text = str(segment or "").strip()
    if not text or "[" not in text or "]" not in text:
        return None

    predicate = text[text.find("[") + 1 : text.rfind("]")]
    if not predicate:
        return None

    # Prefer explicit element-based qualifier used by each segment type.
    target_element = f"{segment_id}01"
    if segment_id == "L11":
        target_element = "L1102"
    elif segment_id == "LIN":
        target_element = "LIN02"

    explicit_match = re.search(
        rf"(?i)\b{re.escape(target_element)}\b\s*=\s*['\"]?([A-Za-z0-9]{{1,10}})['\"]?",
        predicate,
    )
    if explicit_match:
        return explicit_match.group(1).strip().upper()

    # LIN can carry multiple qualifier/value pairs (LIN02/LIN04/LIN06...).
    # Do not infer LIN02 from a generic fallback like LIN06="PO".
    if segment_id == "LIN":
        return None

    # Fallback: first short code-like predicate value.
    fallback_match = re.search(r"=\s*['\"]?([A-Za-z0-9]{1,4})['\"]?", predicate)
    if fallback_match:
        return fallback_match.group(1).strip().upper()

    return None


def extract_datetime_token(text: str) -> dt.datetime | None:
    value = str(text or "").strip()
    if not value:
        return None

    patterns = [
        r"(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:Z|[+-]\d{2}:?\d{2}))",
        r"(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})",
        r"(\d{4}-\d{2}-\d{2})",
        r"(\d{4}/\d{2}/\d{2})",
        r"(\d{2}/\d{2}/\d{4})",
    ]

    for pattern in patterns:
        match = re.search(pattern, value)
        if not match:
            continue
        token = match.group(1)
        try:
            if "T" in token:
                cleaned = token.replace("Z", "+00:00")
                if re.search(r"[+-]\d{4}$", cleaned):
                    cleaned = cleaned[:-5] + cleaned[-5:-2] + ":" + cleaned[-2:]
                return dt.datetime.fromisoformat(cleaned)
            if "/" in token:
                if token.count("/") == 2 and len(token.split("/")[0]) == 4:
                    return dt.datetime.strptime(token, "%Y/%m/%d")
                if token.count("/") == 2:
                    return dt.datetime.strptime(token, "%m/%d/%Y")
            if "-" in token:
                return dt.datetime.strptime(token, "%Y-%m-%d")
        except ValueError:
            continue

    return None


def parse_value(raw_value: str | None) -> str:
    text = "" if raw_value is None else str(raw_value).replace("\r", "").strip()
    if not text:
        return "SAMPLE_VALUE"

    first_line = text.split("\n", 1)[0].strip()
    if not first_line:
        return "SAMPLE_VALUE"

    hardcode_match = re.search(r"(?i)hardcode\s*(?:=|to)?\s*['\"]?([^'\"]+)['\"]?", first_line)
    if hardcode_match:
        return hardcode_match.group(1).strip()

    if re.search(r"(?i)continuous\s+sequential\s+number", first_line):
        return "1"

    if re.search(r"(?i)\bmap\s+source\s+to\s+target\b|\bif\s+source\b|\brefer\b|\blookup\b", first_line):
        return "SV"

    if re.search(r"\[[^\]]+\]/[A-Za-z0-9_:-]+$", first_line):
        return "SAMPLE_VALUE"

    if re.search(r"(?i)\b[A-Z0-9]{2,4}\d{2}\b\s*=", first_line):
        return "SAMPLE_VALUE"

    if re.search(r"(?i)\b\d+\s+digits\b", first_line):
        return "SAMPLE_VALUE"

    if re.fullmatch(r"(?i)no\s+mapping", first_line):
        return "S"

    dt_value = extract_datetime_token(first_line)
    if dt_value is not None:
        if re.search(r"(?i)cc?yymmdd|yyyymmdd|yyyy[-/]?mm[-/]?dd", first_line):
            return dt_value.strftime("%Y%m%d")
        if re.search(r"(?i)yymmdd", first_line):
            return dt_value.strftime("%y%m%d")
        if re.search(r"(?i)hhmmss", first_line):
            return dt_value.strftime("%H%M%S")
        if re.search(r"(?i)hhmm", first_line):
            return dt_value.strftime("%H%M")

    if re.search(r"(?i)current\s+date|system\s+date|today", first_line):
        return dt.datetime.now().strftime("%Y%m%d")

    if re.search(r"(?i)current\s+time|system\s+time|time\s+now", first_line):
        return dt.datetime.now().strftime("%H%M")

    if re.search(r"(?i)format|expressed\s+as|for\s+example|if\s+.+\s+then\s+map", first_line):
        return "SAMPLE_VALUE"

    return first_line


def extract_inline_element_assignments(raw_value: str | None, segment_id: str) -> dict[int, str]:
    text = str(raw_value or "").replace("\r", " ")
    if not text or segment_id not in {"N1", "L11", "LIN"}:
        return {}

    assignments: dict[int, str] = {}
    pattern = rf"(?i)\b{re.escape(segment_id)}(\d{{2}})\b\s*=\s*['\"]?([A-Za-z0-9]{{1,10}})['\"]?"
    for pos_text, value in re.findall(pattern, text):
        pos = int(pos_text)
        if pos < 1:
            continue
        assignments[pos] = value.strip().upper()

    return assignments


def extract_path_element_assignments(path: str, segment_id: str) -> dict[int, str]:
    normalized = normalize_path(path)
    if not normalized:
        return {}

    raw_parts = split_segments(normalized)
    segment_part = None
    for part in raw_parts:
        if strip_segment_predicate(local_name(part)).upper() == segment_id:
            segment_part = part
            break

    if not segment_part or "[" not in segment_part or "]" not in segment_part:
        return {}

    predicate = segment_part[segment_part.find("[") + 1 : segment_part.rfind("]")]
    if not predicate:
        return {}

    assignments: dict[int, str] = {}
    pattern = rf"(?i)\b{re.escape(segment_id)}(\d{{2}})\b\s*=\s*['\"]?([A-Za-z0-9]{{1,20}})['\"]?"
    for pos_text, value in re.findall(pattern, predicate):
        pos = int(pos_text)
        if pos < 1:
            continue
        assignments[pos] = value.strip().upper()

    return assignments


def detect_path_column(rows: list[tuple], start_row: int = 0) -> int | None:
    sample_rows = rows[start_row : start_row + 500]
    if not sample_rows:
        return None

    max_cols = max((len(r) for r in sample_rows), default=0)
    best_col: int | None = None
    best_hits = 0
    best_ratio = 0.0

    for col_idx in range(max_cols):
        non_empty = 0
        hits = 0
        for row in sample_rows:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            non_empty += 1
            if looks_like_x12_path(text):
                hits += 1

        if non_empty == 0:
            continue

        ratio = hits / non_empty
        if hits > best_hits or (hits == best_hits and ratio > best_ratio):
            best_col = col_idx
            best_hits = hits
            best_ratio = ratio

    if best_col is None:
        return None

    if best_hits < 1 or best_ratio < 0.2:
        return None

    return best_col


def detect_best_worksheet(workbook) -> str:
    best_sheet = workbook.sheetnames[0]
    best_hits = -1

    for name in workbook.sheetnames:
        sheet = workbook[name]
        rows = list(sheet.iter_rows(values_only=True))
        col_idx = detect_path_column(rows, 0)
        if col_idx is None:
            continue

        hits = 0
        for row in rows:
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None and looks_like_x12_path(str(value)):
                hits += 1

        if hits > best_hits:
            best_hits = hits
            best_sheet = name

    return best_sheet


def load_x12_rows(
    file_path: Path,
    worksheet_name: str | None = None,
    path_column_ref: str | None = None,
    value_column_ref: str | None = None,
) -> tuple[list[tuple[str, str | None]], str, int]:
    workbook = load_workbook(filename=file_path, data_only=True)

    if worksheet_name:
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{worksheet_name}' not found in workbook.")
        sheet_name = worksheet_name
    else:
        sheet_name = detect_best_worksheet(workbook)

    sheet = workbook[sheet_name]
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return [], sheet_name, 0

    if path_column_ref:
        path_idx = excel_col_to_index(path_column_ref)
    else:
        path_idx = detect_path_column(all_rows, 0)
        if path_idx is None:
            raise ValueError("Could not auto-detect an X12 path column. Try --path-column-ref B")

    value_idx = excel_col_to_index(value_column_ref) if value_column_ref else None

    loaded: list[tuple[str, str | None]] = []
    for row in all_rows:
        path_cell = row[path_idx] if path_idx < len(row) else None
        if path_cell is None:
            continue
        path_text = str(path_cell).strip()
        if not path_text:
            continue

        # Some mapping sheets store multiple X12 paths in a single cell (newline-separated).
        # Expand them so each path is processed independently.
        path_candidates = [line.strip() for line in path_text.splitlines() if line.strip()]
        if not path_candidates:
            continue

        value_text = None
        if value_idx is not None and value_idx < len(row):
            cell_value = row[value_idx]
            value_text = None if cell_value is None else str(cell_value)
        else:
            for candidate_idx in range(path_idx + 1, min(len(row), path_idx + 4)):
                cell_value = row[candidate_idx]
                if cell_value is None:
                    continue
                candidate_text = str(cell_value).strip()
                if not candidate_text:
                    continue
                value_text = candidate_text
                break

            # If nothing was found on the right, allow qualifier hints from nearby left cells.
            if value_text is None:
                for candidate_idx in range(max(0, path_idx - 3), path_idx):
                    cell_value = row[candidate_idx]
                    if cell_value is None:
                        continue
                    candidate_text = str(cell_value).strip()
                    if not candidate_text:
                        continue
                    if re.search(r"(?i)\b(?:N101|L1101|L1102)\b\s*=", candidate_text):
                        value_text = candidate_text
                        break

        for candidate_path in path_candidates:
            if not looks_like_x12_path(candidate_path):
                continue
            loaded.append((candidate_path, value_text))

    return loaded, sheet_name, path_idx


def parse_path_to_segment(path: str) -> tuple[str, int, str, str | None] | None:
    normalized = normalize_path(path)
    if not normalized:
        return None

    raw_parts = split_segments(normalized)
    parts = [strip_segment_predicate(local_name(p)) for p in raw_parts]
    if len(parts) < 3:
        return None

    if parts[0].upper() != "X12":
        return None

    idx = 1
    if idx < len(parts) and re.fullmatch(r"TS_\d{3}", parts[idx], flags=re.IGNORECASE):
        idx += 1

    while idx < len(parts) and re.fullmatch(r"GROUP_\d+", parts[idx], flags=re.IGNORECASE):
        idx += 1

    if idx + 1 >= len(parts):
        return None

    segment_id = parts[idx].upper()
    if not re.fullmatch(r"[A-Z0-9]{2,3}", segment_id):
        return None

    element_name = parts[idx + 1]
    pos_match = re.search(r"(\d{2})$", element_name)
    if not pos_match:
        return None
    position = int(pos_match.group(1))
    if position < 1:
        return None

    def _occ_part(raw_part: str, seg_raw: str | None = None) -> str:
        stripped = strip_segment_predicate(local_name(raw_part))
        if re.fullmatch(r"GROUP_\d+", stripped, flags=re.IGNORECASE):
            # Case 1: predicate on GROUP itself contains N1/N101 qualifier
            if "[" in raw_part:
                pred = raw_part[raw_part.find("[") + 1 : raw_part.rfind("]")]
                n101_m = re.search(r'N1\s*/\s*N101\s*=\s*["\']?([A-Za-z0-9]{1,10})["\']?', pred, re.IGNORECASE)
                if n101_m:
                    return f"{stripped}[{n101_m.group(1).upper()}]"
            # Case 2: GROUP_4 has no predicate, but child N1 has N101 qualifier
            # Look ahead at the segment part itself if it's N1
            if seg_raw is not None:
                child_stripped = strip_segment_predicate(local_name(seg_raw))
                if child_stripped.upper() == "N1" and "[" in seg_raw:
                    spred = seg_raw[seg_raw.find("[") + 1 : seg_raw.rfind("]")]
                    n101_m2 = re.search(r'N101\s*=\s*["\']?([A-Za-z0-9]{1,10})["\']?', spred, re.IGNORECASE)
                    if n101_m2:
                        return f"{stripped}[{n101_m2.group(1).upper()}]"
        return stripped

    # Pass the segment raw part (raw_parts[idx]) to the last GROUP before the segment
    seg_raw_part = raw_parts[idx]
    occurrence_parts = []
    for i, part in enumerate(raw_parts[: idx + 1]):
        stripped = strip_segment_predicate(local_name(part))
        if re.fullmatch(r"GROUP_\d+", stripped, flags=re.IGNORECASE) and i == idx - 1:
            # This GROUP is the direct parent of the segment — pass segment raw for case 2
            occurrence_parts.append(_occ_part(part, seg_raw_part))
        else:
            occurrence_parts.append(_occ_part(part))
    occurrence_key = "/".join(occurrence_parts)
    qualifier = None
    if segment_id in {"N1", "L11", "REF", "DTM", "LIN"}:
        qualifier = extract_segment_qualifier(raw_parts[idx], segment_id)

    return segment_id, position, occurrence_key, qualifier


def get_sample_value(segment_id: str, position: int, transaction_set: str) -> str:
    now = dt.datetime.now()

    if segment_id == "ISA":
        defaults = {
            1: "00",
            2: "          ",
            3: "00",
            4: "          ",
            5: "ZZ",
            6: "SENDERID      ",
            7: "ZZ",
            8: "RECEIVERID    ",
            9: now.strftime("%y%m%d"),
            10: now.strftime("%H%M"),
            11: "U",
            12: "00401",
            13: "000000001",
            14: "0",
            15: "T",
            16: ":",
        }
        return defaults.get(position, "SAMPLE_VALUE")

    if segment_id == "GS":
        gs01 = "QM" if transaction_set == "214" else ("SH" if transaction_set == "856" else "OW")
        defaults = {
            1: gs01,
            2: "SENDER",
            3: "RECEIVER",
            4: now.strftime("%Y%m%d"),
            5: now.strftime("%H%M"),
            6: "0001",
            7: "X",
            8: "004010",
        }
        return defaults.get(position, "SAMPLE_VALUE")

    if segment_id == "ST":
        return transaction_set if position == 1 else ("0001" if position == 2 else "SAMPLE_VALUE")

    if segment_id == "SE":
        return "0000" if position == 1 else ("0001" if position == 2 else "SAMPLE_VALUE")

    if segment_id == "GE":
        return "1" if position in (1, 2) else "SAMPLE_VALUE"

    if segment_id == "IEA":
        return "1" if position == 1 else ("000000001" if position == 2 else "SAMPLE_VALUE")

    if segment_id == "N1":
        defaults = {
            1: "SH",
            2: "SAMPLE_VALUE",
            3: "93",
            4: "SAMPLE_VALUE",
        }
        return defaults.get(position, "SAMPLE_VALUE")

    if segment_id == "AT7":
        defaults = {
            1: "X2",
            2: "NS",
            3: "",
            4: "",
            5: now.strftime("%Y%m%d"),
            6: now.strftime("%H%M"),
            7: "UT",
        }
        return defaults.get(position, "")

    if segment_id == "MS1":
        defaults = {
            1: "CITY",
            2: "ST",
            3: "US",
            4: "",
            5: "",
        }
        return defaults.get(position, "")

    if segment_id == "AT8":
        defaults = {
            1: "G",
            2: "G",
            3: "1",
            4: "1",
            5: "1",
            6: "1",
            7: "KG",
        }
        return defaults.get(position, "1")

    return "SAMPLE_VALUE"


def format_edi_value(segment_id: str, position: int, value: str) -> str:
    text = str(value or "")

    if segment_id == "ISA" and position in (6, 8):
        return text[:15].ljust(15, " ")

    if segment_id == "ISA" and position == 9:
        if re.fullmatch(r"\d{8}", text):
            return text[2:]
        if re.fullmatch(r"\d{6}", text):
            return text

    if segment_id == "ISA" and position == 10:
        if re.fullmatch(r"\d{4}", text):
            return text

    if segment_id == "ISA" and position == 13:
        digits = re.sub(r"\D", "", text)
        if not digits:
            return "000000001"
        return digits[-9:].zfill(9)

    return text


def fit_value_to_element_rule(value: str, rule: SefElementDefinition) -> str:
    text = str(value or "")
    if text == "":
        return ""

    max_length = int(rule.max_length or 0)
    min_length = int(rule.min_length or 0)
    data_type = (rule.data_type or "").upper()

    def build_short_placeholder() -> str:
        if data_type == "N":
            target_length = max(1, min(max_length or 1, min_length or 1))
            return "1" * target_length

        if min_length > 0 and max_length > 0:
            target_length = min_length if min_length <= max_length else max_length
        elif min_length > 0:
            target_length = min_length
        elif max_length > 0:
            target_length = min(max_length, 10)
        else:
            target_length = 6

        target_length = max(1, target_length)
        return "X" * target_length

    if text in {"SAMPLE_VALUE", "SAMPLE_V", "SAMPLE", "SAM", "SAMP"}:
        return build_short_placeholder()

    if max_length > 0 and len(text) > max_length:
        if data_type == "N":
            digits = re.sub(r"\D", "", text) or "0"
            text = digits[:max_length]
        else:
            text = text[:max_length]

    if min_length > 0 and len(text) < min_length:
        pad_char = "0" if data_type == "N" else "X"
        text = text.ljust(min_length, pad_char)

    return text


def normalize_edi_to_sef(
    edi_text: str,
    segment_definitions: dict[str, SefSegmentDefinition],
    element_definitions: dict[str, SefElementDefinition],
    transaction_set: str,
) -> str:
    normalized_segments: list[str] = []
    for segment_id, elements in parse_edi_segments(edi_text):
        if segment_id not in segment_definitions:
            normalized_segments.append(f"{segment_id}*{'*'.join(elements)}~")
            continue

        definition = segment_definitions[segment_id]
        working_values = list(elements)
        required_max = max(definition.required_positions, default=0)
        while len(working_values) < required_max:
            pos = len(working_values) + 1
            working_values.append(get_sample_value(segment_id, pos, transaction_set))

        normalized_values: list[str] = []
        for pos, value in enumerate(working_values, start=1):
            text = str(value or "")
            if pos - 1 < len(definition.element_ids):
                element_id = definition.element_ids[pos - 1]
                rule = element_definitions.get(element_id)
                if rule is not None:
                    text = fit_value_to_element_rule(text, rule)
            normalized_values.append(text)

        normalized_segments.append(f"{segment_id}*{'*'.join(normalized_values)}~")

    return "\n".join(normalized_segments)


def dedupe_consecutive_lines(lines: list[str]) -> list[str]:
    deduped: list[str] = []
    last = None
    for line in lines:
        text = str(line or "").strip()
        if not text:
            continue
        if text == last:
            continue
        deduped.append(text)
        last = text
    return deduped


def collapse_duplicate_group1_n1_blocks(lines: list[str]) -> list[str]:
    group1_child_ids = {"N2", "N3", "N4", "G61", "G62", "L11"}
    seen_qualifiers: set[str] = set()
    collapsed: list[str] = []
    skip_duplicate_block = False
    placeholder_qualifiers = {"", "X", "XX", "XXX", "S", "SV", "SAMPLE", "SAMPLE_VALUE"}

    for line in lines:
        segment_text = str(line or "").strip()
        if not segment_text:
            continue

        segment_id = segment_text.split("*", 1)[0].upper()
        if segment_id == "N1":
            parts = segment_text.rstrip("~").split("*")
            qualifier = parts[1].strip().upper() if len(parts) > 1 else ""
            skip_duplicate_block = qualifier in placeholder_qualifiers or (qualifier in seen_qualifiers and qualifier != "")
            if skip_duplicate_block:
                continue
            if qualifier:
                seen_qualifiers.add(qualifier)
            collapsed.append(segment_text)
            continue

        if skip_duplicate_block:
            if segment_id in group1_child_ids:
                continue
            skip_duplicate_block = False

        collapsed.append(segment_text)

    return collapsed


def drop_placeholder_only_at8(lines: list[str]) -> list[str]:
    has_rich_at8 = False
    for line in lines:
        text = str(line or "").strip()
        if not text.startswith("AT8*"):
            continue
        parts = text.rstrip("~").split("*")
        if len(parts) > 2:
            has_rich_at8 = True
            break

    if not has_rich_at8:
        return lines

    placeholder_values = {"", "X", "XX", "XXX", "S", "SV", "SAMPLE", "SAMPLE_VALUE"}
    kept: list[str] = []
    for line in lines:
        text = str(line or "").strip()
        if text.startswith("AT8*"):
            parts = text.rstrip("~").split("*")
            value1 = parts[1].strip().upper() if len(parts) > 1 else ""
            if len(parts) == 2 and value1 in placeholder_values:
                continue
        kept.append(text)
    return kept


def collapse_to_single_at8(lines: list[str]) -> list[str]:
    at8_indices = [idx for idx, line in enumerate(lines) if str(line or "").strip().startswith("AT8*")]
    if len(at8_indices) <= 1:
        return lines

    def at8_rank(line: str) -> tuple[int, int]:
        parts = str(line or "").strip().rstrip("~").split("*")
        at801 = parts[1].strip().upper() if len(parts) > 1 else ""
        # Prefer explicit AT801=G, then keep the most complete AT8 line.
        return (1 if at801 == "G" else 0, len(parts))

    best_index = max(at8_indices, key=lambda idx: at8_rank(lines[idx]))
    collapsed: list[str] = []
    for idx, line in enumerate(lines):
        text = str(line or "").strip()
        if text.startswith("AT8*") and idx != best_index:
            continue
        collapsed.append(text)
    return collapsed


def collapse_to_single_lx(lines: list[str]) -> list[str]:
    seen_lx = False
    collapsed: list[str] = []
    for line in lines:
        text = str(line or "").strip()
        if text.startswith("LX*"):
            if seen_lx:
                continue
            seen_lx = True
        collapsed.append(text)
    return collapsed


def merge_at7_ms1_groups(lines: list[str]) -> list[str]:
    at7_indices = [idx for idx, line in enumerate(lines) if str(line or "").strip().startswith("AT7*")]
    ms1_indices = [idx for idx, line in enumerate(lines) if str(line or "").strip().startswith("MS1*")]
    if len(at7_indices) <= 1 and len(ms1_indices) <= 1:
        return lines

    placeholder_values = {
        "",
        "X",
        "XX",
        "XXX",
        "S",
        "SV",
        "SAMPLE",
        "SAMPLE_VALUE",
        "XXXX",
        "XXXXXXX",
        "XXXXXXXX",
    }

    def rank_line(line: str) -> tuple[int, int]:
        parts = str(line or "").strip().rstrip("~").split("*")[1:]
        meaningful = sum(1 for part in parts if str(part or "").strip().upper() not in placeholder_values)
        return (meaningful, len(parts))

    best_at7 = lines[max(at7_indices, key=lambda idx: rank_line(lines[idx]))] if at7_indices else None
    best_ms1 = lines[max(ms1_indices, key=lambda idx: rank_line(lines[idx]))] if ms1_indices else None

    anchor_candidates = at7_indices + ms1_indices
    anchor = min(anchor_candidates) if anchor_candidates else len(lines)

    merged: list[str] = []
    inserted = False
    for idx, line in enumerate(lines):
        text = str(line or "").strip()
        if idx == anchor and not inserted:
            if best_at7 is not None:
                merged.append(str(best_at7 or "").strip())
            if best_ms1 is not None:
                merged.append(str(best_ms1 or "").strip())
            inserted = True

        if text.startswith("AT7*") or text.startswith("MS1*"):
            continue
        merged.append(text)

    if not inserted:
        if best_at7 is not None:
            merged.append(str(best_at7 or "").strip())
        if best_ms1 is not None:
            merged.append(str(best_ms1 or "").strip())

    return merged


def drop_placeholder_only_cd3(lines: list[str]) -> list[str]:
    placeholder_values = {
        "",
        "X",
        "XX",
        "XXX",
        "S",
        "SV",
        "SAMPLE",
        "SAMPLE_VALUE",
        "XXXX",
        "XXXXXXX",
        "XXXXXXXX",
    }

    kept: list[str] = []
    for line in lines:
        text = str(line or "").strip()
        if text.startswith("CD3*"):
            parts = text.rstrip("~").split("*")[1:]
            meaningful = any(str(part or "").strip().upper() not in placeholder_values for part in parts)
            if not meaningful:
                continue
        kept.append(text)
    return kept


def drop_placeholder_only_n1_children(lines: list[str]) -> list[str]:
    # Keep N2/N3/N4 only when they carry at least one meaningful mapped value.
    child_ids = set()
    placeholder_values = {
        "",
        "X",
        "XX",
        "XXX",
        "S",
        "SV",
        "SAMPLE",
        "SAMPLE_VALUE",
        "XXXX",
        "XXXXXXX",
        "XXXXXXXX",
    }

    kept: list[str] = []
    for line in lines:
        text = str(line or "").strip()
        segment_id = text.split("*", 1)[0].upper() if text else ""
        if segment_id in child_ids:
            parts = text.rstrip("~").split("*")[1:]
            meaningful_count = sum(
                1 for part in parts if str(part or "").strip().upper() not in placeholder_values
            )
            min_meaningful = 2 if segment_id == "N4" else 1
            if meaningful_count < min_meaningful:
                continue
        kept.append(text)

    return kept


def parse_occurrence_key_parts(occurrence_key: str) -> tuple[list[str], int]:
    text = re.sub(r"\[[^\]]*\]", "", str(occurrence_key or "").strip())
    parts = [part for part in text.split("/") if part]
    if len(parts) >= 2 and parts[0].upper() == "X12" and re.fullmatch(r"TS_\d{3}", parts[1], flags=re.IGNORECASE):
        parts = parts[2:]

    repeat_index = 0
    if parts and "#" in parts[-1]:
        base_part, repeat_text = parts[-1].rsplit("#", 1)
        parts[-1] = base_part
        if repeat_text.isdigit():
            repeat_index = int(repeat_text)

    return parts, repeat_index


def build_generic_path_sort_key(parts: list[str], segment_id: str, repeat_index: int, occurrence_key: str) -> tuple:
    part_keys: list[tuple[int, int | str]] = []
    for part in parts:
        group_match = re.fullmatch(r"GROUP_(\d+)", str(part or ""), flags=re.IGNORECASE)
        if group_match:
            part_keys.append((0, int(group_match.group(1))))
        else:
            part_keys.append((1, str(part or "").upper()))

    return (tuple(part_keys), repeat_index, segment_id, occurrence_key)


def build_occurrence_sort_key(item: dict, transaction_set: str) -> tuple:
    occurrence_key = str(item.get("occurrence_key", ""))
    segment_id = str(item.get("segment_id", "")).upper()
    parts, repeat_index = parse_occurrence_key_parts(occurrence_key)

    if transaction_set == "856":
        # Top-level 856 order: BSN=10, DTM=20, then GROUP_1 loops
        top856 = {"BSN": 10, "DTM": 20, "GROUP_1": 30}
        # Inside GROUP_1[HL03="S"] (shipment): HL emitted automatically, then REF/DTM/N1
        g1s_order = {"REF": 20, "DTM": 30, "GROUP_4": 40}
        # Inside GROUP_1[HL03="I"] (item): LIN, SN1, TD3, REF, DTM, GROUP_4
        g1i_order = {"LIN": 10, "SN1": 20, "TD3": 25, "REF": 30, "DTM": 40, "GROUP_4": 60}
        # Inside GROUP_4 (N1 loop)
        g4_order = {"N1": 10, "N2": 20, "N3": 30, "N4": 40, "REF": 50}

        if not parts:
            return (999, repeat_index, segment_id, occurrence_key)

        first = parts[0].upper()
        if first != "GROUP_1":
            return (0, top856.get(first, 900), repeat_index, segment_id, occurrence_key)

        # Determine HL type from occurrence key predicate
        hl_type = "S"
        hl_m = re.search(r"HL03\s*=\s*[\"']?([A-Z])[\"']?", occurrence_key, re.IGNORECASE)
        if hl_m:
            hl_type = hl_m.group(1).upper()

        hl_order = 0 if hl_type == "S" else 1

        if len(parts) == 1:
            return (1, hl_order, repeat_index, segment_id, occurrence_key)

        second = parts[1].upper()
        if second.startswith("GROUP_4"):
            # Extract N1 qualifier from raw occurrence_key (predicates preserved there)
            n1q_m = re.search(r'GROUP_4\[([A-Z0-9]{1,10})\]', occurrence_key, re.IGNORECASE)
            n1_q = n1q_m.group(1).upper() if n1q_m else ""
            inner_order = g4_order.get(segment_id, 900)
            return (1, hl_order, repeat_index, g1s_order.get("GROUP_4", 40) if hl_type == "S" else g1i_order.get("GROUP_4", 60), n1_q, inner_order, occurrence_key)

        order_map = g1s_order if hl_type == "S" else g1i_order
        return (1, hl_order, repeat_index, order_map.get(segment_id, 900), occurrence_key)

    if transaction_set != "214":
        return build_generic_path_sort_key(parts, segment_id, repeat_index, occurrence_key)

    top_level_order = {
        "B10": 10,
        "L11": 20,
        "MAN": 30,
        "K1": 40,
        "GROUP_1": 50,
        "MS3": 60,
        "GROUP_2": 70,
    }
    group1_order = {
        "N1": 10,
        "N2": 20,
        "N3": 30,
        "N4": 40,
        "G61": 50,
        "G62": 60,
        "L11": 70,
    }
    group2_order = {
        "GROUP_3": 10,
        "L11": 20,
        "MAN": 30,
        "Q7": 40,
        "K1": 50,
        "AT5": 60,
        "AT8": 70,
        "GROUP_4": 80,
    }
    group3_order = {
        "AT7": 10,
        "MS1": 20,
        "MS2": 30,
    }
    group4_order = {
        "CD3": 10,
        "L11": 20,
        "GROUP_5": 30,
        "NM1": 40,
        "Q7": 50,
        "AT8": 60,
        "MAN": 70,
        "GROUP_6": 80,
    }

    if not parts:
        return (999, repeat_index, segment_id, occurrence_key)

    first_part = parts[0].upper()
    if first_part != "GROUP_1" and first_part != "GROUP_2":
        return (0, top_level_order.get(first_part, 900), repeat_index, segment_id, occurrence_key)

    if first_part == "GROUP_1":
        return (1, repeat_index, group1_order.get(segment_id, 900), segment_id, occurrence_key)

    if len(parts) == 1:
        return (2, 0, repeat_index, segment_id, occurrence_key)

    second_part = parts[1].upper()
    if second_part == "GROUP_3":
        return (2, group2_order["GROUP_3"], repeat_index, group3_order.get(segment_id, 900), segment_id, occurrence_key)

    if second_part == "GROUP_4":
        return (2, group2_order["GROUP_4"], repeat_index, group4_order.get(segment_id, 900), segment_id, occurrence_key)

    return (2, group2_order.get(segment_id, 900), repeat_index, segment_id, occurrence_key)


def build_sample_edi(rows: list[tuple[str, str | None]], transaction_set: str) -> str:
    control_maps: dict[str, dict[int, str]] = {}
    occurrences: OrderedDict[str, dict] = OrderedDict()
    occurrence_counters: dict[str, int] = {}
    latest_occurrence_by_base: dict[str, str] = {}
    line_item_count = 0

    for path, raw_value in rows:
        parsed = parse_path_to_segment(path)
        if not parsed:
            continue

        segment_id, position, occurrence_key, qualifier = parsed
        inline_assignments = extract_path_element_assignments(path, segment_id)
        raw_inline_assignments = extract_inline_element_assignments(raw_value, segment_id)
        for assigned_pos, assigned_value in raw_inline_assignments.items():
            inline_assignments.setdefault(assigned_pos, assigned_value)

        if segment_id == "AT8":
            at8_match = re.search(r"(?i)\bAT801\b\s*=\s*['\"]?([A-Za-z0-9]{1,10})['\"]?", str(path or ""))
            if at8_match:
                inline_assignments.setdefault(1, at8_match.group(1).strip().upper())

        value = parse_value(raw_value)
        if value == NO_MAPPING_TOKEN:
            value = ""

        if segment_id in {"N1", "L11", "REF", "DTM", "LIN"} and qualifier:
            if segment_id in {"N1", "L11", "REF", "DTM"}:
                occurrence_key = f"{occurrence_key}[{qualifier}]"
            if segment_id == "N1":
                # Always ensure position 1 gets the qualifier value (e.g. LW, SU, ST)
                inline_assignments.setdefault(1, qualifier)
            if segment_id == "L11":
                inline_assignments.setdefault(2, qualifier)
            if segment_id == "REF":
                # REF01 is the qualifier element
                inline_assignments.setdefault(1, qualifier)
            if segment_id == "DTM":
                # DTM01 is the qualifier element
                inline_assignments.setdefault(1, qualifier)
            if segment_id == "LIN":
                # LIN02 is the first product/service ID qualifier.
                inline_assignments.setdefault(2, qualifier)

        assigned_for_position = inline_assignments.get(position)
        if assigned_for_position and value in {"", "S", "SV", "SAM", "SAMPLE", "SAMPLE_VALUE"}:
            value = assigned_for_position

        if segment_id == "REF" and position != 1 and value in {"SV", "S", "SAMPLE"}:
            value = "SAMPLE_VALUE"

        if segment_id in CONTROL_SEGMENTS:
            control_maps.setdefault(segment_id, {})
            control_maps[segment_id].setdefault(position, value)
            continue

        if segment_id == "CTT":
            continue

        # Determine 856 HL type from the raw path predicate (GROUP_1[HL/HL03="S"] etc.)
        hl_type_856 = None
        hl_m_856 = re.search(r"GROUP_1\s*\[.*?HL03\s*=\s*[\"']?([A-Z])[\"']?", path, re.IGNORECASE)
        if hl_m_856:
            hl_type_856 = hl_m_856.group(1).upper()

        base_occurrence_key = occurrence_key
        # Keep S and I loop occurrences distinct to prevent cross-HL value bleed.
        if transaction_set == "856" and hl_type_856 and "/GROUP_1/" in base_occurrence_key:
            base_occurrence_key = f"{base_occurrence_key}[HL03={hl_type_856}]"

        resolved_occurrence_key = latest_occurrence_by_base.get(base_occurrence_key, base_occurrence_key)

        existing = occurrences.get(resolved_occurrence_key)
        starts_new_repetition = position == 1
        if segment_id == "LIN":
            # Consolidate LIN details into one LIN segment per loop.
            starts_new_repetition = False
        if not starts_new_repetition and existing and existing["elements"]:
            earlier_assignments = {
                assigned_pos: assigned_value
                for assigned_pos, assigned_value in inline_assignments.items()
                if assigned_pos < position
            }
            has_conflicting_assignment = any(
                assigned_pos in existing["elements"] and existing["elements"][assigned_pos] != assigned_value
                for assigned_pos, assigned_value in earlier_assignments.items()
            )
            repeats_current_position = position in existing["elements"]
            starts_new_repetition = has_conflicting_assignment or repeats_current_position

        # A repeated earliest element for the same base path means a new segment repetition.
        if starts_new_repetition:
            if existing and existing["elements"]:
                next_count = occurrence_counters.get(base_occurrence_key, 0) + 1
                occurrence_counters[base_occurrence_key] = next_count
                resolved_occurrence_key = f"{base_occurrence_key}#{next_count}"

        latest_occurrence_by_base[base_occurrence_key] = resolved_occurrence_key

        if resolved_occurrence_key not in occurrences:
            occurrences[resolved_occurrence_key] = {
                "segment_id": segment_id,
                "occurrence_key": resolved_occurrence_key,
                "elements": {},
                "hl_type": hl_type_856,
            }
        elif hl_type_856 and not occurrences[resolved_occurrence_key].get("hl_type"):
            occurrences[resolved_occurrence_key]["hl_type"] = hl_type_856

        for assigned_pos, assigned_value in inline_assignments.items():
            occurrences[resolved_occurrence_key]["elements"].setdefault(assigned_pos, assigned_value)

        occurrences[resolved_occurrence_key]["elements"].setdefault(position, value)

    output: list[str] = []
    data_lines: list[str] = []
    group3_lines_by_loop: OrderedDict[str, list[str]] = OrderedDict()
    group3_marker = "__GROUP3_BLOCK__"

    def emit_segment(segment_id: str, source_map: dict[int, str] | None, force_count: int = 0) -> str:
        placeholder_values = {
            "",
            "X",
            "XX",
            "XXX",
            "S",
            "SV",
            "SAMPLE",
            "SAMPLE_VALUE",
            "XXXX",
            "XXXXXXX",
            "XXXXXXXX",
        }
        today = dt.datetime.now().strftime("%Y%m%d")
        clock = dt.datetime.now().strftime("%H%M")

        if segment_id == "ISA":
            values = [
                format_edi_value("ISA", pos, get_sample_value("ISA", pos, transaction_set))
                for pos in range(1, 17)
            ]
            return f"ISA*{'*'.join(values)}~"

        if segment_id == "GS":
            values = [
                format_edi_value("GS", pos, get_sample_value("GS", pos, transaction_set))
                for pos in range(1, 9)
            ]
            return f"GS*{'*'.join(values)}~"

        source = source_map or {}
        max_pos = max([*source.keys(), force_count] or [1])
        values: list[str] = []
        for pos in range(1, max_pos + 1):
            raw = source.get(pos, "")
            if not raw or raw == "SAMPLE_VALUE":
                raw = get_sample_value(segment_id, pos, transaction_set)

            if segment_id == "G61":
                if pos == 1 and str(raw).strip().upper() in placeholder_values:
                    raw = "CN"
                elif pos == 2 and str(raw).strip().upper() in placeholder_values:
                    raw = today
                elif pos > 2 and str(raw).strip().upper() in placeholder_values:
                    raw = ""

            if segment_id == "G62":
                if pos == 1 and str(raw).strip().upper() in placeholder_values:
                    raw = "86"
                elif pos == 2 and str(raw).strip().upper() in placeholder_values:
                    raw = today
                elif pos == 3 and str(raw).strip().upper() in placeholder_values:
                    raw = "8"
                elif pos == 4 and str(raw).strip().upper() in placeholder_values:
                    raw = clock
                elif pos > 4 and str(raw).strip().upper() in placeholder_values:
                    raw = ""

            if segment_id == "DTM":
                raw_norm = str(raw).strip().upper()
                if pos == 2 and (raw_norm in placeholder_values or raw_norm.startswith("SAMPLE")):
                    raw = today

            if segment_id == "AT8":
                if pos == 6:
                    value_text = str(raw).strip().upper()
                    if value_text not in {"B", "C", "D", "E", "F", "G", "H", "L", "M", "N", "R", "S", "T", "U", "V", "X"}:
                        raw = "L"
                elif pos == 7:
                    value_text = str(raw).strip()
                    if not re.fullmatch(r"\d+(?:\.\d+)?", value_text):
                        raw = "1"

            values.append(format_edi_value(segment_id, pos, raw))

        if segment_id == "AT7" and len(values) >= 4:
            if values[0] and values[2]:
                values[2] = ""
                values[3] = ""

        if segment_id == "MS1" and len(values) >= 5:
            if values[0] and values[3]:
                values[3] = ""
                values[4] = ""

        if segment_id == "DTM":
            # Keep DTM01/DTM02, but do not force optional trailing DTM elements with placeholders.
            while len(values) > 2 and values[-1] in placeholder_values:
                values.pop()

        return f"{segment_id}*{'*'.join(values)}~"

    output.append(emit_segment("ISA", control_maps.get("ISA"), 16))
    output.append(emit_segment("GS", control_maps.get("GS"), 8))
    output.append(emit_segment("ST", control_maps.get("ST"), 2))

    data_segment_count = 0
    b10_emitted = False
    lx_counter = 0
    ordered_occurrences = sorted(
        occurrences.values(),
        key=lambda item: build_occurrence_sort_key(item, transaction_set),
    )

    for item in ordered_occurrences:
        segment_id = item["segment_id"]
        occurrence_key = item.get("occurrence_key", "")
        if segment_id in CONTROL_SEGMENTS:
            continue

        if transaction_set == "214" and segment_id == "B10":
            if b10_emitted:
                continue
            b10_emitted = True

        emitted_line = emit_segment(segment_id, item["elements"], 0)

        if transaction_set == "214":
            group3_match = re.search(r"^(.*?/GROUP_3)/[^/#]+(?:#(\d+))?$", occurrence_key)
            if group3_match:
                group3_path = group3_match.group(1)
                loop_index = int(group3_match.group(2) or 0)
                group3_loop_key = f"{group3_path}#{loop_index}"
                if group3_marker not in data_lines:
                    data_lines.append(group3_marker)
                group3_lines_by_loop.setdefault(group3_loop_key, [])
                group3_lines_by_loop[group3_loop_key].append(emitted_line)
            else:
                group3_loop_key = None

            if group3_loop_key is not None:
                if segment_id == "LIN":
                    line_item_count += 1
                data_segment_count += 1
                continue

        data_lines.append(emitted_line)
        if segment_id == "LIN":
            line_item_count += 1
        data_segment_count += 1

    # For 856, inject HL segments before each GROUP_1 loop block.
    # Extract HL type from the original raw paths (predicate is stripped in occurrence keys).
    if transaction_set == "856":
        # Determine which segments belong to each HL type (from raw path predicates).
        # Build a precise (occurrence_key) -> hl_type map from stored occurrences.
        hl_types_seen: list[str] = []
        hl_types_set: set[str] = set()
        for path_raw, _ in rows:
            g1_m = re.search(r"GROUP_1\s*\[.*?HL03\s*=\s*[\"']?([A-Z])[\"']?", path_raw, re.IGNORECASE)
            if g1_m:
                ht = g1_m.group(1).upper()
                if ht not in hl_types_set:
                    hl_types_seen.append(ht)
                    hl_types_set.add(ht)

        n1_children = {"N2", "N3", "N4"}  # must follow their N1 parent

        # Build HL segment lines
        hl_type_to_line: dict[str, str] = {}
        hl_counter = 0
        shipment_hl_num: int | None = None
        for ht in hl_types_seen:
            hl_counter += 1
            if ht == "S":
                shipment_hl_num = hl_counter
                hl_type_to_line[ht] = f"HL*{hl_counter}**S~"
            else:
                parent = str(shipment_hl_num) if shipment_hl_num else "1"
                hl_type_to_line[ht] = f"HL*{hl_counter}*{parent}*{ht}~"

        # Build a map from emitted_line -> hl_type using occurrence hl_type field.
        # We match by occurrence_key stored in ordered_occurrences.
        line_to_hl: dict[int, str] = {}  # index in data_lines -> hl_type
        ship_ht = next((h for h in hl_types_seen if h == "S"), hl_types_seen[0] if hl_types_seen else None)
        item_ht = next((h for h in hl_types_seen if h != "S"), ship_ht)

        # Re-emit in order to map each data_line to its hl_type
        data_line_hl: list[str | None] = []  # parallel to data_lines
        for item in ordered_occurrences:
            seg = item["segment_id"]
            if seg in CONTROL_SEGMENTS or seg == "CTT":
                continue
            ht = item.get("hl_type")
            # For segments with no HL type (BSN, DTM at top level) ht stays None
            data_line_hl.append(ht)

        # Classify data_lines into top-level and per-HL buckets using data_line_hl.
        top_segs_856 = {"BSN", "DTM"}
        top_lines: list[str] = []
        hl_buckets: dict[str, list[str]] = {ht: [] for ht in hl_types_seen}
        last_n1_bucket: str | None = None

        for idx_dl, dl in enumerate(data_lines):
            seg = dl.rstrip("~").split("*")[0].upper()
            ht = data_line_hl[idx_dl] if idx_dl < len(data_line_hl) else None

            if seg in top_segs_856 and ht is None:
                top_lines.append(dl)
            elif seg in n1_children:
                target = last_n1_bucket if last_n1_bucket else ship_ht
                if target and target in hl_buckets:
                    hl_buckets[target].append(dl)
                else:
                    top_lines.append(dl)
            else:
                # Use hl_type from occurrence; fallback to shipment
                bucket = ht if (ht and ht in hl_buckets) else ship_ht
                if bucket and bucket in hl_buckets:
                    if seg == "N1":
                        last_n1_bucket = bucket
                    hl_buckets[bucket].append(dl)
                else:
                    top_lines.append(dl)

        # Business rule override for this 856 mapping:
        # HL-S should contain DTM*193/194 and HL-I should contain DTM*161.
        if ship_ht and ship_ht in hl_buckets:
            ship_bucket = hl_buckets[ship_ht]
            item_bucket = hl_buckets.get(item_ht, []) if item_ht else []

            def dtm_qualifier(line: str) -> str | None:
                parts = line.rstrip("~").split("*")
                if len(parts) >= 2 and parts[0].upper() == "DTM":
                    return parts[1].strip().upper()
                return None

            # Collect all DTM lines from both buckets keyed by qualifier.
            dtm_pool: dict[str, list[str]] = {}
            for src_line in [*ship_bucket, *item_bucket]:
                q = dtm_qualifier(src_line)
                if not q:
                    continue
                dtm_pool.setdefault(q, []).append(src_line)

            ROUTED_QUALIFIERS = {"193", "194", "161"}

            def rebuild_bucket(base_lines: list[str], allowed_qualifiers: list[str]) -> list[str]:
                non_dtm = [
                    line
                    for line in base_lines
                    if line.rstrip("~").split("*")[0].upper() != "DTM"
                ]

                # Preserve DTMs whose qualifier is NOT being rerouted (keep them in place)
                local_dtm = [
                    line for line in base_lines
                    if line.rstrip("~").split("*")[0].upper() == "DTM"
                    and (dtm_qualifier(line) or "") not in ROUTED_QUALIFIERS
                ]

                # Add routed qualifiers from the pool
                kept_dtm: list[str] = []
                for q in allowed_qualifiers:
                    if q in dtm_pool and dtm_pool[q]:
                        kept_dtm.append(dtm_pool[q][0])

                insert_idx = next(
                    (idx for idx, line in enumerate(non_dtm) if line.rstrip("~").split("*")[0].upper() == "N1"),
                    len(non_dtm),
                )
                return non_dtm[:insert_idx] + local_dtm + kept_dtm + non_dtm[insert_idx:]

            hl_buckets[ship_ht] = rebuild_bucket(ship_bucket, ["193", "194"])
            if item_ht and item_ht in hl_buckets:
                hl_buckets[item_ht] = rebuild_bucket(item_bucket, ["161"])

        def ensure_ref_qualifier(bucket_lines: list[str], qualifier: str, value: str = "SAMPLE_VALUE") -> list[str]:
            q = str(qualifier or "").strip().upper()
            if not q:
                return bucket_lines

            for line in bucket_lines:
                parts = line.rstrip("~").split("*")
                if len(parts) >= 3 and parts[0].upper() == "REF" and parts[1].strip().upper() == q and parts[2].strip():
                    return bucket_lines

            ref_line = f"REF*{q}*{value}~"
            insert_idx = next(
                (idx for idx, line in enumerate(bucket_lines) if line.rstrip("~").split("*")[0].upper() == "N1"),
                len(bucket_lines),
            )
            return bucket_lines[:insert_idx] + [ref_line] + bucket_lines[insert_idx:]

        if ship_ht and ship_ht in hl_buckets:
            hl_buckets[ship_ht] = ensure_ref_qualifier(hl_buckets[ship_ht], "CN")
        if item_ht and item_ht in hl_buckets:
            hl_buckets[item_ht] = ensure_ref_qualifier(hl_buckets[item_ht], "SCA")

        # Reassemble: top lines, then each HL block prefixed with its HL segment
        expanded_856: list[str] = []
        expanded_856.extend(top_lines)
        for ht in hl_types_seen:
            if ht in hl_type_to_line:
                expanded_856.append(hl_type_to_line[ht])
            expanded_856.extend(hl_buckets[ht])
        data_lines = expanded_856

    if transaction_set == "214" and group3_lines_by_loop:
        grouped_group3_lines: list[str] = []
        for loop_lines in group3_lines_by_loop.values():
            lx_counter += 1
            grouped_group3_lines.append(emit_segment("LX", {1: str(lx_counter)}, 1))
            grouped_group3_lines.extend(loop_lines)
            data_segment_count += 1

        expanded_lines: list[str] = []
        for line in data_lines:
            if line == group3_marker:
                expanded_lines.extend(grouped_group3_lines)
            else:
                expanded_lines.append(line)
        data_lines = expanded_lines

    if transaction_set in {"214", "856"}:
        data_lines = collapse_duplicate_group1_n1_blocks(data_lines)
        data_lines = drop_placeholder_only_n1_children(data_lines)
        data_lines = drop_placeholder_only_at8(data_lines)
        data_lines = collapse_to_single_at8(data_lines)
        data_lines = collapse_to_single_lx(data_lines)
        data_lines = merge_at7_ms1_groups(data_lines)
        data_lines = drop_placeholder_only_cd3(data_lines)

    if transaction_set == "214":
        first_k1_index = next((idx for idx, line in enumerate(data_lines) if line.startswith("K1*")), None)
        if first_k1_index is not None:
            l11_after_k1 = [line for line in data_lines[first_k1_index + 1 :] if line.startswith("L11*")]
            if l11_after_k1:
                kept_tail = [line for line in data_lines[first_k1_index + 1 :] if not line.startswith("L11*")]
                data_lines = data_lines[:first_k1_index] + l11_after_k1 + [data_lines[first_k1_index]] + kept_tail

    data_segment_count = len(data_lines)

    output.extend(data_lines)

    se_count = (len(data_lines) + 2) if transaction_set == "214" else (3 + data_segment_count)
    if transaction_set != "214":
        output.append(f"CTT*{line_item_count}~")
    output.append(f"SE*{se_count}*0001~")
    output.append("GE*1*0001~")
    output.append("IEA*1*000000001~")

    output = dedupe_consecutive_lines(output)
    return "\n".join(output)


def parse_sef_set_tokens(set_expression: str) -> list[SefToken]:
    tokens: list[SefToken] = []
    text = str(set_expression or "")
    depth = 0
    idx = 0

    while idx < len(text):
        ch = text[idx]
        if ch == "{":
            depth += 1
            idx += 1
            continue
        if ch == "}":
            depth = max(0, depth - 1)
            idx += 1
            continue
        if ch != "[" or depth > 0:
            idx += 1
            continue

        end = text.find("]", idx)
        if end < 0:
            break

        content = text[idx + 1 : end]
        fields = [part.strip() for part in content.split(",")]
        segment_field = fields[0] if fields else ""
        segment_id = segment_field.split("@", 1)[0].strip().upper()
        if re.fullmatch(r"[A-Z0-9]{2,3}", segment_id):
            usage = (fields[1] if len(fields) > 1 else "").strip().upper()
            max_use_text = (fields[2] if len(fields) > 2 else "").strip()
            if max_use_text == ">1":
                max_use = 10**9
            elif max_use_text.isdigit():
                max_use = int(max_use_text)
            else:
                max_use = 1
            tokens.append(SefToken(segment_id, usage == "M", max_use))

        idx = end + 1

    return tokens


def parse_sef_segment_definitions(lines: list[str]) -> dict[str, SefSegmentDefinition]:
    definitions: dict[str, SefSegmentDefinition] = {}
    for line in lines:
        text = str(line or "").strip()
        if not text or "=" not in text:
            continue

        segment_id, body = text.split("=", 1)
        segment_id = segment_id.strip().upper()
        if not re.fullmatch(r"[A-Z0-9]{2,3}", segment_id):
            continue

        matches = re.findall(r"\[([^\]]+)\]", body)
        required_positions: list[int] = []
        element_ids: list[str] = []
        for idx, match in enumerate(matches, start=1):
            fields = [part.strip() for part in match.split(",")]
            element_id = fields[0] if fields else ""
            usage = (fields[1] if len(fields) > 1 else "").upper()
            element_ids.append(element_id)
            if usage == "M":
                required_positions.append(idx)

        definitions[segment_id] = SefSegmentDefinition(
            segment_id=segment_id,
            element_count=len(matches),
            required_positions=required_positions,
            element_ids=element_ids,
        )

    return definitions


def parse_sef_element_definitions(lines: list[str]) -> dict[str, SefElementDefinition]:
    definitions: dict[str, SefElementDefinition] = {}
    for line in lines:
        text = str(line or "").strip()
        match = re.match(r"^(\d+)=([A-Z]{1,3}),(\d+),(\d+)\s*$", text)
        if not match:
            continue

        element_id = match.group(1)
        if element_id in definitions:
            continue

        definitions[element_id] = SefElementDefinition(
            element_id=element_id,
            data_type=match.group(2),
            min_length=int(match.group(3)),
            max_length=int(match.group(4)),
        )

    return definitions


def load_sef_schema(sef_path: Path, transaction_set: str) -> tuple[list[SefToken], dict[str, SefSegmentDefinition], dict[str, SefElementDefinition]] | None:
    text = sef_path.read_text(encoding="utf-8", errors="ignore")
    lines = text.splitlines()

    set_expression = ""
    segment_lines: list[str] = []
    element_lines: list[str] = []
    in_sets = False
    in_segs = False
    in_elms = False

    for raw_line in lines:
        line = str(raw_line or "")
        if line.startswith(".SETS"):
            in_sets, in_segs, in_elms = True, False, False
            continue
        if line.startswith(".SEGS"):
            in_sets, in_segs, in_elms = False, True, False
            continue
        if line.startswith(".ELMS"):
            in_sets, in_segs, in_elms = False, False, True
            continue
        if re.match(r"^\.[A-Z]", line):
            in_sets = in_segs = in_elms = False
            continue

        if in_sets:
            match = re.match(r"^(\d{3})=(.*)$", line)
            if match and match.group(1) == transaction_set:
                set_expression = match.group(2)
            continue

        if in_segs:
            segment_lines.append(line)
            continue

        if in_elms:
            element_lines.append(line)

    if not set_expression.strip():
        return None

    return (
        parse_sef_set_tokens(set_expression),
        parse_sef_segment_definitions(segment_lines),
        parse_sef_element_definitions(element_lines),
    )


def parse_edi_segments(edi_text: str) -> list[tuple[str, list[str]]]:
    segments: list[tuple[str, list[str]]] = []
    for raw_segment in re.split(r"~\s*", str(edi_text or "")):
        segment_text = raw_segment.strip()
        if not segment_text:
            continue
        parts = segment_text.split("*")
        segment_id = str(parts[0] or "").strip().upper()
        if not re.fullmatch(r"[A-Z0-9]{2,3}", segment_id):
            continue
        segments.append((segment_id, parts[1:]))
    return segments


def validate_edi_against_sef(
    edi_text: str,
    schema: tuple[list[SefToken], dict[str, SefSegmentDefinition], dict[str, SefElementDefinition]],
) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    set_tokens, segment_definitions, element_definitions = schema
    segments = parse_edi_segments(edi_text)
    if not segments:
        return ["Generated EDI contains no readable segments."], warnings

    token_counts: dict[int, int] = {}
    pattern_index = 0
    top_level_ids = {token.segment_id for token in set_tokens}

    for line_no, (segment_id, elements) in enumerate(segments, start=1):
        if segment_id in segment_definitions:
            definition = segment_definitions[segment_id]
            if len(elements) > definition.element_count:
                errors.append(
                    f"Segment {segment_id} at position {line_no} has {len(elements)} elements, exceeding SEF definition of {definition.element_count}."
                )

            for required_pos in definition.required_positions:
                value = elements[required_pos - 1] if len(elements) >= required_pos else ""
                if value == "":
                    errors.append(
                        f"Segment {segment_id} at position {line_no} is missing required element {segment_id}{required_pos:02d}."
                    )

            for pos, value in enumerate(elements, start=1):
                if pos - 1 >= len(definition.element_ids):
                    continue
                element_id = definition.element_ids[pos - 1]
                if not element_id or element_id not in element_definitions:
                    continue
                if not value:
                    continue

                rule = element_definitions[element_id]
                actual_len = len(value)
                if rule.max_length > 0 and actual_len > rule.max_length:
                    errors.append(
                        f"{segment_id}{pos:02d}: element too long, actual: {actual_len}, standard: {rule.max_length}"
                    )
                elif rule.min_length > 0 and actual_len < rule.min_length:
                    errors.append(
                        f"{segment_id}{pos:02d}: element too short, actual: {actual_len}, standard: {rule.min_length}"
                    )
        else:
            warnings.append(f"Segment {segment_id} at position {line_no} is not defined in the loaded SEF.")

        if segment_id not in top_level_ids:
            continue

        matched = False
        for candidate_index in range(pattern_index, len(set_tokens)):
            candidate = set_tokens[candidate_index]
            if candidate.segment_id != segment_id:
                continue

            for skipped_index in range(pattern_index, candidate_index):
                skipped = set_tokens[skipped_index]
                skipped_count = token_counts.get(skipped_index, 0)
                if skipped.is_required and skipped_count == 0:
                    warnings.append(
                        f"Segment {segment_id} at position {line_no} appears before required segment {skipped.segment_id} in the SEF sequence."
                    )

            current_count = token_counts.get(candidate_index, 0)
            if current_count >= candidate.max_use:
                continue

            token_counts[candidate_index] = current_count + 1
            pattern_index = candidate_index
            matched = True
            break

        if not matched:
            warnings.append(f"Segment {segment_id} at position {line_no} does not match expected SEF sequence.")

    for idx, token in enumerate(set_tokens):
        if token.is_required and token_counts.get(idx, 0) == 0:
            errors.append(f"Required segment {token.segment_id} is missing according to the SEF definition.")

    return errors, warnings


def detect_transaction_set(rows: list[tuple[str, str | None]], fallback: str) -> str:
    for path, _ in rows:
        match = re.search(r"/TS_(\d{3})/", normalize_path(path), flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return fallback


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a sample X12 EDI file from Excel path/value mapping rows."
    )
    parser.add_argument("input", help="Path to .xlsx mapping file")
    parser.add_argument(
        "-o",
        "--output",
        default="sample.edi",
        help="Output EDI file path (default: sample.edi)",
    )
    parser.add_argument(
        "--worksheet",
        default=None,
        help="Optional worksheet name (default: auto-detect best worksheet)",
    )
    parser.add_argument(
        "--path-column-ref",
        default=None,
        help="Use a fixed Excel column for X12 paths (e.g. B or 2)",
    )
    parser.add_argument(
        "--value-column-ref",
        default=None,
        help="Optional Excel column for mapped values/guidance (e.g. C or 3)",
    )
    parser.add_argument(
        "--transaction-set",
        default="214",
        help="Fallback X12 transaction set id, e.g. 214/856/861 (default: 214)",
    )
    parser.add_argument(
        "--sef-file",
        default=None,
        help="Optional SEF file path used to validate the generated EDI",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if input_path.suffix.lower() != ".xlsx":
        print("Input must be an .xlsx file.", file=sys.stderr)
        return 1

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1

    try:
        rows, worksheet_name, path_idx = load_x12_rows(
            file_path=input_path,
            worksheet_name=args.worksheet,
            path_column_ref=args.path_column_ref,
            value_column_ref=args.value_column_ref,
        )
    except Exception as exc:
        print(f"Failed to read Excel file: {exc}", file=sys.stderr)
        return 1

    if not rows:
        print("No usable X12 path rows found in the Excel file.", file=sys.stderr)
        return 1

    transaction_set = detect_transaction_set(rows, fallback=str(args.transaction_set).strip())
    edi_text = build_sample_edi(rows, transaction_set)

    validation_errors: list[str] = []
    validation_warnings: list[str] = []
    schema = None
    if args.sef_file:
        sef_path = Path(args.sef_file).expanduser().resolve()
        if not sef_path.exists():
            print(f"SEF file not found: {sef_path}", file=sys.stderr)
            return 1
        schema = load_sef_schema(sef_path, transaction_set)
        if schema is None:
            validation_warnings.append(f"No matching transaction set {transaction_set} found in SEF: {sef_path}")
        else:
            _, segment_definitions, element_definitions = schema
            edi_text = normalize_edi_to_sef(edi_text, segment_definitions, element_definitions, transaction_set)
            edi_text = "\n".join(dedupe_consecutive_lines(edi_text.splitlines()))
            validation_errors, validation_warnings = validate_edi_against_sef(edi_text, schema)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(edi_text + "\n", encoding="utf-8")

    print(f"Generated EDI file: {output_path}")
    print(f"Worksheet used: {worksheet_name}")
    print(f"Path column used: {index_to_excel_col(path_idx)} ({path_idx + 1})")
    print(f"Transaction set: {transaction_set}")
    print(f"Rows processed: {len(rows)}")
    if args.sef_file:
        print(f"SEF file: {Path(args.sef_file).expanduser().resolve()}")
        print(f"Validation errors: {len(validation_errors)}")
        print(f"Validation warnings: {len(validation_warnings)}")
        for message in validation_errors[:10]:
            print(f"ERROR: {message}")
        for message in validation_warnings[:10]:
            print(f"WARNING: {message}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
